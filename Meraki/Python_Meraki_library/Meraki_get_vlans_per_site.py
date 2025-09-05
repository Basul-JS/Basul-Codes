# Created on 2025-09-04 by JS
    # Script goes through entire org and outputs VLAN associated with each site/ network 
# 2025-09-05 Updated to increase parallelism 
# please ensure the below modules are installed 
    # pip install meraki
    # pip install openpyxl

from __future__ import annotations

import logging
import re
import sys
import time
import threading
from collections import deque
from datetime import datetime
from getpass import getpass
from typing import Any, Dict, Iterable, List, Optional, Set, Tuple, TypedDict, cast
from concurrent.futures import ThreadPoolExecutor, as_completed
import csv

import meraki
from meraki.exceptions import APIError

# ---------- Optional Excel dependency ----------
try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet
    from openpyxl.worksheet.filters import AutoFilter
    HAS_OPENPYXL: bool = True
except Exception:
    HAS_OPENPYXL = False
    Workbook = object        # type: ignore[assignment]
    Worksheet = object       # type: ignore[assignment]
    AutoFilter = object      # type: ignore[assignment]
    get_column_letter = None # type: ignore[assignment]

# ---------------- Logging ----------------
timestamp: str = datetime.now().strftime("%Y%m%d_%H%M%S")
logging.basicConfig(
    filename=f"meraki_script_{timestamp}.log",
    level=logging.DEBUG,
    format="%(asctime)s - %(levelname)s - %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)
# (no console handler -> keep console clean)

# ---------------- Constants ----------------
MAX_RETRIES: int = 5
MAX_WORKERS: int = 24          # good balance for ~1000 sites; limiter governs RPS
MERAKI_RPS: int = 5            # conservative org limit
EXCLUDED_VLANS: Set[int] = {100, 110, 210, 220, 230, 235, 240}

# ---------------- Types ----------------
class VLANRow(TypedDict):
    organizationName: str
    networkId: str
    networkName: str
    networkProductTypes: str
    vlanId: str
    vlanName: str
    subnet: str
    applianceIp: str
    dhcpHandling: str

# ---------------- Rate limiter ----------------
class RateLimiter:
    def __init__(self, max_per_sec: int) -> None:
        self.max_per_sec = max_per_sec
        self._lock = threading.Lock()
        self._calls: deque[float] = deque()

    def acquire(self) -> None:
        window = 1.0
        while True:
            with self._lock:
                now = time.monotonic()
                # drop timestamps older than 1s
                while self._calls and (now - self._calls[0]) > window:
                    self._calls.popleft()
                if len(self._calls) < self.max_per_sec:
                    self._calls.append(now)
                    logger.debug("RateLimiter token at %.3f", now)
                    return
                sleep_time = window - (now - self._calls[0])
                logger.debug("RateLimiter sleeping %.3fs", sleep_time)
            time.sleep(sleep_time if sleep_time > 0 else 0.01)

limiter = RateLimiter(MERAKI_RPS)

# ---------------- Auth + Dashboard ----------------
API_KEY: str = getpass("Enter your API key (input hidden): ")
dashboard: meraki.DashboardAPI = meraki.DashboardAPI(
    API_KEY,
    suppress_logging=True,
    wait_on_rate_limit=True,   # SDK honors Retry-After on 429
    retry_4xx_error=True,
    maximum_retries=10,
    single_request_timeout=60,
)

def select_org() -> Tuple[str, str]:
    orgs: List[Dict[str, Any]] = dashboard.organizations.getOrganizations()
    if not orgs:
        logger.error("No organisations returned from API")
        print("No organisations available for this API key.")
        sys.exit(1)

    print("Organizations:")
    for idx, org in enumerate(orgs, 1):
        name = cast(str, org.get("name", ""))
        oid = cast(str, org.get("id", ""))
        print(f"{idx}. {name} (ID: {oid})")

    raw: str = input("Select organization by number (or press Enter to cancel): ").strip()
    if not raw:
        logger.error("No organisation selected by user")
        print("No organisation selected.")
        sys.exit(1)

    try:
        org_idx = int(raw)
        if org_idx < 1 or org_idx > len(orgs):
            raise ValueError("out of range")
    except Exception:
        logger.error("Invalid organisation selection: %s", raw)
        print("Invalid selection.")
        sys.exit(1)

    chosen: Dict[str, Any] = orgs[org_idx - 1]
    org_id = cast(str, chosen.get("id", ""))
    org_name = cast(str, chosen.get("name", ""))
    logger.debug("Selected org %s (%s)", org_name, org_id)
    return org_id, org_name

org_id, org_name = select_org()
name_filter: str = input("Filter networks by name (partial, optional): ").strip().lower()
only_with_appliance: bool = True  # set False to include all networks

# ---------------- Helpers ----------------
def try_int(value: Any) -> Optional[int]:
    try:
        return int(value)
    except Exception:
        return None

def api_request_with_retries(func: Any, *args: Any, **kwargs: Any) -> Any:
    fname: str = getattr(func, "__name__", str(func))  # ensure bound for except paths
    for i in range(MAX_RETRIES):
        limiter.acquire()
        try:
            logger.debug("API call %s attempt %d args=%s kwargs=%s", fname, i + 1, args, kwargs)
            return func(*args, **kwargs)
        except APIError as e:
            status: Optional[int] = getattr(e, "status", None)
            msg: str = getattr(e, "message", str(e))
            retry_after: Optional[float] = None
            ra_any: Any = getattr(e, "retry_after", None)
            try:
                if ra_any is not None:
                    retry_after = float(ra_any)
            except Exception:
                retry_after = None

            if status == 429:
                wait = retry_after if (retry_after and retry_after > 0) else float(2 ** i)
                logger.warning("429 on %s. Sleeping %.2fs (retry %d/%d)", fname, wait, i + 1, MAX_RETRIES)
                time.sleep(wait)
                continue

            logger.error("APIError in %s: status=%s msg=%s", fname, status, msg)
            raise
        except Exception:
            logger.exception("Unexpected error in %s", fname)
            raise
    raise RuntimeError(f"Max retries exceeded for {fname}")

def list_org_networks(org_id_param: str) -> List[Dict[str, Any]]:
    nets: List[Dict[str, Any]] = api_request_with_retries(
        dashboard.organizations.getOrganizationNetworks,
        org_id_param,
        total_pages="all",
        perPage=1000,  # reduce pagination overhead for ~1000 sites
    )
    if name_filter:
        nets = [n for n in nets if name_filter in cast(str, n.get("name", "")).lower()]
    if only_with_appliance:
        nets = [n for n in nets if "appliance" in cast(List[str], n.get("productTypes") or [])]
    logger.debug("Networks after filter: %d", len(nets))
    return nets

def fetch_network_vlan_rows(network: Dict[str, Any]) -> List[VLANRow]:
    """
    Two-step (explicit):
      1) getNetworkApplianceVlansSettings -> skip early if vlansEnabled=False
      2) getNetworkApplianceVlans       -> enumerate VLANs
    """
    rows: List[VLANRow] = []
    net_id: str = cast(str, network.get("id", ""))
    net_name: str = cast(str, network.get("name", ""))
    product_types_list: List[str] = cast(List[str], network.get("productTypes") or [])
    net_types: str = ",".join(product_types_list)

    if not net_id:
        return rows

    try:
        # Step 1: VLANs enabled?
        settings: Dict[str, Any] = api_request_with_retries(
            dashboard.appliance.getNetworkApplianceVlansSettings, net_id
        )
        if not (settings and settings.get("vlansEnabled", False)):
            logger.debug("VLANs disabled for %s (%s)", net_name, net_id)
            return rows

        # Step 2: Fetch VLANs
        vlans: List[Dict[str, Any]] = api_request_with_retries(
            dashboard.appliance.getNetworkApplianceVlans, net_id
        )
        for v in vlans or []:
            raw_id: Any = v.get("id", "")
            vlan_id_int: Optional[int] = try_int(raw_id)
            if vlan_id_int is not None and vlan_id_int in EXCLUDED_VLANS:
                logger.debug("Skipping excluded VLAN %s on %s", raw_id, net_name)
                continue

            row: VLANRow = VLANRow(
                organizationName=org_name,
                networkId=net_id,
                networkName=net_name,
                networkProductTypes=net_types,
                vlanId=str(raw_id),
                vlanName=cast(str, v.get("name", "")),
                subnet=cast(str, v.get("subnet", "")),
                applianceIp=cast(str, v.get("applianceIp", "")),
                dhcpHandling=cast(str, v.get("dhcpHandling", "")),
            )
            rows.append(row)
        logger.debug("Fetched %d VLAN rows for %s", len(rows), net_name)

    except APIError as e:
        status = getattr(e, "status", None)
        # If settings call failed (e.g., no appliance), treat as "no VLANs"
        if status in (400, 404):
            logger.debug("No VLANs/settings for %s (%s) [status %s]", net_name, net_id, status)
            return rows
        logger.exception("APIError fetching VLANs for %s (%s)", net_name, net_id)
    except Exception:
        logger.exception("Unexpected error fetching VLANs for %s", net_name)

    return rows

def sanitize_filename_part(text: str) -> str:
    cleaned = re.sub(r"[^\w\s\-\._]", "", text).strip()
    cleaned = re.sub(r"\s+", "_", cleaned)
    return cleaned or "org"

# ---------- Tabular writers (Excel preferred, CSV/TSV fallback) ----------
def data_has_commas(rows: List[VLANRow], keys: List[str]) -> bool:
    for r in rows:
        for k in keys:
            val = str(r.get(k, ""))
            if "," in val:
                return True
    return False

def write_csv_or_tsv(filepath_no_ext: str, rows: List[VLANRow], fieldnames: List[str]) -> str:
    use_tsv: bool = data_has_commas(rows, fieldnames)
    delimiter: str = "\t" if use_tsv else ","
    ext: str = "tsv" if use_tsv else "csv"
    out_path: str = f"{filepath_no_ext}.{ext}"

    with open(out_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, delimiter=delimiter, quoting=csv.QUOTE_MINIMAL)
        writer.writeheader()
        for r in rows:
            writer.writerow({k: r.get(k, "") for k in fieldnames})

    logger.info("Wrote %s (%d rows)", out_path, len(rows))
    return out_path

def write_excel(out_path: str, rows: List[VLANRow], fieldnames: List[str]) -> None:
    wb: Workbook = Workbook()  # type: ignore[call-arg]
    ws_opt = getattr(wb, "active", None)
    if ws_opt is None:
        ws_opt = wb.create_sheet(title="VLANs")  # type: ignore[operator]
    ws: Worksheet = cast(Worksheet, ws_opt)      # type: ignore[assignment]
    ws.title = "VLANs"

    ws.append(fieldnames)
    for r in rows:
        ws.append([r.get(k, "") for k in fieldnames])

    ws.freeze_panes = "A2"
    if get_column_letter is not None:
        last_col_letter = get_column_letter(len(fieldnames))  # type: ignore[operator]
        last_row = ws.max_row
        ws.auto_filter = AutoFilter(ref=f"A1:{last_col_letter}{last_row}")  # type: ignore[call-arg]

        col_widths: List[int] = [len(h) for h in fieldnames]
        for row_vals in ws.iter_rows(min_row=2, max_row=last_row, max_col=len(fieldnames), values_only=True):
            row_iter: Iterable[Optional[Any]] = cast(Iterable[Optional[Any]], row_vals)
            for i, val in enumerate(row_iter):
                l = len(str(val)) if val is not None else 0
                if l > col_widths[i]:
                    col_widths[i] = l
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = min(max(w + 2, 10), 60)  # type: ignore[operator]

    wb.save(out_path)
    logger.info("Excel written: %s", out_path)

def write_tabular(org_name_val: str, rows: List[VLANRow], fieldnames: List[str]) -> str:
    org_part: str = sanitize_filename_part(org_name_val)
    base_path: str = f"org_vlans_{org_part}_{timestamp}"

    if HAS_OPENPYXL:
        out_xlsx = f"{base_path}.xlsx"
        try:
            write_excel(out_xlsx, rows, fieldnames)
            return out_xlsx
        except Exception:
            logger.exception("Excel write failed; falling back to CSV/TSV")
            return write_csv_or_tsv(base_path, rows, fieldnames)
    else:
        logger.debug("openpyxl not available; writing CSV/TSV")
        return write_csv_or_tsv(base_path, rows, fieldnames)

def sort_key_vlan(row: VLANRow) -> Tuple[str, Tuple[int, str]]:
    vlan_int = try_int(row["vlanId"])
    return (row["networkName"], (vlan_int if vlan_int is not None else 10**9, row["vlanId"]))

# ---------------- Main ----------------
def main() -> None:
    networks: List[Dict[str, Any]] = list_org_networks(org_id)
    if not networks:
        print("No matching networks found in this organization.")
        return

    print(f"Scanning {len(networks)} network(s)...")  # minimal console
    logger.debug("ThreadPool workers: %d", MAX_WORKERS)

    all_rows: List[VLANRow] = []
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        future_map = {executor.submit(fetch_network_vlan_rows, n): n for n in networks}
        for fut in as_completed(future_map):
            n = future_map[fut]
            try:
                rows = fut.result()
                all_rows.extend(rows)
            except Exception:
                logger.exception("Worker failed for %s", n.get("name", "<unknown>"))

    all_rows.sort(key=sort_key_vlan)

    fieldnames: List[str] = [
        "organizationName",
        "networkId",
        "networkName",
        "networkProductTypes",
        "vlanId",
        "vlanName",
        "subnet",
        "applianceIp",
        "dhcpHandling",
    ]

    out_file: str = write_tabular(org_name, all_rows, fieldnames)

    nets_with_vlans: Set[str] = {r["networkId"] for r in all_rows}
    print(
        f"\n✅ Completed. Networks scanned: {len(networks)} | "
        f"Networks with VLANs (after excludes): {len(nets_with_vlans)} | Rows: {len(all_rows)}"
    )
    print(f"Output: {out_file}")
    logger.info("Completed run: networks=%d rows=%d file=%s", len(networks), len(all_rows), out_file)

if __name__ == "__main__":
    main()

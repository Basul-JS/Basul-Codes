# DRY-RUN MERAKI SNAPSHOT + SIMULATED PLAN (NO CHANGES MADE)
# Adds: Production-like template selection, simulated profile assignment,
#       simulated port-override reapply, and a triage sheet for MS with no compatible profile.
#
# Dependencies:
#   pip install meraki openpyxl
#
# Exports:
#   <network>_pre_<ts>.xlsx              (PRE_Snapshot)
#   <network>_post_simulated_<ts>.xlsx   (POST_Simulated, with Planned_Actions, Port_Overrides_Sim, MS_No_Compatible_Profile)
#   <network>_diff_<ts>.xlsx             (Diff - VLAN-focused)
#
# This script NEVER changes Meraki state.

import meraki
import logging
import re
import json
from datetime import datetime
from getpass import getpass
import sys
from typing import Any, Dict, List, Optional, Tuple, Set, cast
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ---------------- Config ----------------
BASE_URL: str = "https://api.meraki.com/api/v1"
timestamp: str = datetime.now().strftime("%Y%m%d_%H%M%S")

logging.basicConfig(
    filename=f"meraki_dryrun_plan_{timestamp}.log",
    level=logging.DEBUG,
    format="%(asctime)s - %(levelname)s - %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S"
)

# ---------------- API Key ----------------
def validate_api_key(key: str) -> bool:
    return bool(re.fullmatch(r"[A-Fa-f0-9]{40}", key or ""))

API_KEY: Optional[str] = None
for _ in range(4):
    API_KEY = getpass("Enter your Meraki API key (hidden): ")
    if validate_api_key(API_KEY):
        break
else:
    print("❌ Invalid API key after 4 attempts")
    sys.exit(1)

dashboard = meraki.DashboardAPI(
    api_key=API_KEY,
    base_url=BASE_URL,
    output_log=False,
    print_console=False,
    suppress_logging=True,
    maximum_retries=5,
    wait_on_rate_limit=True,
)

# ---------------- Helpers: networks/devices/VLANs (read-only) ----------------
def _slug_filename(s: str) -> str:
    s2 = re.sub(r"[^A-Za-z0-9._-]+", "-", s).strip("-_")
    return s2[:80]

def get_all_networks(org_id: str) -> List[Dict[str, Any]]:
    """
    Fetch all networks via manual pagination (startingAfter), avoiding total_pages='all'
    so Pylance remains happy. Also dedupes by id for safety.
    """
    results: List[Dict[str, Any]] = []
    seen_ids: Set[str] = set()
    starting_after: Optional[str] = None

    # Cast the SDK method to Any to avoid false-positive param warnings in some stubs.
    _get_org_nets = cast(Any, dashboard.organizations.getOrganizationNetworks)

    while True:
        try:
            if starting_after:
                page: List[Dict[str, Any]] = _get_org_nets(
                    org_id, perPage=1000, startingAfter=starting_after
                )
            else:
                page = _get_org_nets(org_id, perPage=1000)
        except Exception:
            logging.exception("get_all_networks: page fetch failed")
            break

        if not page:
            break

        for n in page:
            nid = str(n.get("id", "") or "")
            if nid and nid not in seen_ids:
                results.append(n)
                seen_ids.add(nid)

        # If we got fewer than perPage, we've reached the end.
        if len(page) < 1000:
            break

        # Prepare the cursor for the next page.
        starting_after = str(page[-1].get("id", "") or "") or None
        if not starting_after:
            break

    return results


def select_org() -> str:
    orgs = dashboard.organizations.getOrganizations()
    if not orgs:
        print("❌ No organisations found")
        sys.exit(1)

    for i, org in enumerate(orgs, 1):
        print(f"{i}. {org['name']} (ID: {org['id']})")

    raw = input("Select organisation by number: ").strip()
    if not raw.isdigit() or not (1 <= int(raw) <= len(orgs)):
        print("❌ Invalid selection")
        sys.exit(1)
    return cast(str, orgs[int(raw) - 1]["id"])

from typing import Tuple, List, Dict, Any

def fetch_matching_networks(org_id: str, partial: str) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    """
    Return two lists:
      - matches: networks where partial is a substring (case-insensitive)
      - suggestions: networks starting with the same first letter(s) if no matches
    """
    nets: List[Dict[str, Any]] = get_all_networks(org_id)
    partial_lower = partial.lower()

    matches = [n for n in nets if partial_lower in (n.get("name", "")).lower()]

    suggestions: List[Dict[str, Any]] = []
    if not matches and partial_lower:
        # crude "did you mean" → prefix suggestion
        suggestions = [n for n in nets if (n.get("name") or "").lower().startswith(partial_lower[:2])]

    return matches, suggestions


def select_network_interactive(org_id: str) -> Tuple[str, str]:
    """
    Prompt interactively to select a network by searching with a partial name.
    Supports substring search, suggestions, and multi-choice lists.
    """
    while True:
        partial = input("Enter partial network name to search: ").strip()
        if not partial:
            print("No network selected.")
            sys.exit(1)

        matches, suggestions = fetch_matching_networks(org_id, partial)

        if matches:
            if len(matches) == 1:
                m = matches[0]
                print(f"Selected: {m['name']} ({m['id']})")
                return str(m["id"]), str(m["name"])

            print("\nMultiple matches:")
            for i, n in enumerate(matches, 1):
                print(f"{i}. {n['name']} (ID: {n['id']})")

            raw = input("Pick # : ").strip()
            if raw.isdigit() and 1 <= int(raw) <= len(matches):
                chosen = matches[int(raw) - 1]
                print(f"Selected: {chosen['name']} ({chosen['id']})")
                return str(chosen["id"]), str(chosen["name"])
            else:
                print("Invalid choice.")
        else:
            print("No exact/substring matches found.")
            if suggestions:
                print("Did you mean:")
                for n in suggestions:
                    print(f" - {n['name']} (ID: {n['id']})")

        retry = input("Search again? (y/N): ").strip().lower()
        if retry != "y":
            print("No network selected.")
            sys.exit(1)

def fetch_devices(network_id: str) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], List[Dict[str, Any]]]:
    try:
        devices = dashboard.networks.getNetworkDevices(network_id)
    except Exception:
        logging.exception("Failed to fetch devices")
        return [], [], []

    def _mk(d: Dict[str, Any]) -> Dict[str, Any]:
        tags_val = d.get("tags", [])
        tags_list: List[str]
        if isinstance(tags_val, list):
            tags_list = [str(t) for t in tags_val]
        else:
            tags_list = [t for t in str(tags_val or "").split() if t]

        return {
            "serial": d.get("serial", ""),
            "model": d.get("model", ""),
            "tags": tags_list,
            "address": d.get("address", ""),
            "name": d.get("name", ""),
            "switchProfileId": d.get("switchProfileId"),
            "switchProfileName": d.get("switchProfileName"),
        }

    mx = [_mk(d) for d in devices if (d.get("model") or "").startswith("MX")]
    ms = [_mk(d) for d in devices if (d.get("model") or "").startswith("MS")]
    mr = [_mk(d) for d in devices if (d.get("model") or "").upper().startswith(("MR", "CW"))]
    return mx, ms, mr

def fetch_vlan_details(network_id: str) -> List[Dict[str, Any]]:
    try:
        return dashboard.appliance.getNetworkApplianceVlans(network_id)
    except meraki.APIError as e:
        body = getattr(getattr(e, "response", None), "text", "") or ""
        if "VLANs are not enabled" in (getattr(e, "message", "") or "") or "VLANs are not enabled" in body:
            logging.warning("VLANs disabled on this network; returning empty list for dry-run.")
            return []
        logging.exception("APIError while fetching VLANs")
        return []
    except Exception:
        logging.exception("Failed to fetch VLANs")
        return []

def get_warm_spare_status(network_id: str) -> Dict[str, Any]:
    try:
        return dashboard.appliance.getNetworkApplianceWarmSpare(network_id) or {}
    except Exception:
        logging.exception("Failed to read warm spare status")
        return {}

def get_network_template_id(network_id: str) -> Optional[str]:
    try:
        net = dashboard.networks.getNetwork(network_id)
        tpl_id = net.get("configTemplateId")
        return cast(Optional[str], tpl_id)
    except Exception:
        logging.exception("Failed to read network info")
        return None

# ---------------- In-memory "simulation" rules (no changes) ----------------
def _dhcp_is_server(dhcp_handling: Optional[str]) -> bool:
    val = (dhcp_handling or "").strip().lower()
    return val in {"run a dhcp server", "dhcp server", "server", "enabled", "on"}

def simulate_vlan_payloads(vlans: List[Dict[str, Any]]) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    post_vlans: List[Dict[str, Any]] = []
    actions: List[Dict[str, Any]] = []

    for v in vlans:
        vid = v.get("id")
        dhcp = v.get("dhcpHandling")
        would_strip = not _dhcp_is_server(dhcp)

        payload: Dict[str, Any] = {
            "id": vid,
            "name": v.get("name"),
            "subnet": v.get("subnet"),
            "applianceIp": v.get("applianceIp"),
            "dhcpHandling": dhcp,
        }

        if _dhcp_is_server(dhcp):
            if isinstance(v.get("fixedIpAssignments"), dict):
                payload["fixedIpAssignments"] = v.get("fixedIpAssignments")
            if isinstance(v.get("reservedIpRanges"), list):
                payload["reservedIpRanges"] = v.get("reservedIpRanges")

        if would_strip and (v.get("fixedIpAssignments") or v.get("reservedIpRanges")):
            actions.append({
                "scope": "VLAN",
                "vlanId": str(vid),
                "action": "strip_fixed_reserved_for_non_server_dhcp",
                "reason": "Meraki API rejects fixedIpAssignments/reservedIpRanges when DHCP is OFF/RELAY.",
                "before_has_fixed_reserved": True
            })
        else:
            actions.append({
                "scope": "VLAN",
                "vlanId": str(vid),
                "action": "apply_payload",
                "reason": "DHCP is server; assignments/ranges are allowed.",
                "before_has_fixed_reserved": bool(v.get("fixedIpAssignments") or v.get("reservedIpRanges"))
            })

        post_vlans.append(payload)

    return post_vlans, actions

# ---- Port override diffing (read-only) ----
_PORT_FIELDS = [
    "enabled", "name", "tags", "type", "vlan", "voiceVlan", "allowedVlans",
    "poeEnabled", "isolationEnabled", "rstpEnabled", "stpGuard",
    "linkNegotiation", "udld", "accessPolicyType", "accessPolicyNumber",
    "portScheduleId"
]

def _normalize_tags(value: Any) -> List[str]:
    if isinstance(value, list):
        return sorted([str(t) for t in value])
    if isinstance(value, str):
        return sorted([t for t in value.split() if t])
    return []

def _port_dict_by_number(ports: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    out: Dict[str, Dict[str, Any]] = {}
    for p in ports:
        pid = p.get("portId") or p.get("number") or p.get("name")
        if pid is None:
            continue
        out[str(pid)] = p
    return out

def compute_port_overrides(live_ports: List[Dict[str, Any]], tmpl_ports: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    overrides: Dict[str, Dict[str, Any]] = {}
    live = _port_dict_by_number(live_ports)
    tmpl = _port_dict_by_number(tmpl_ports)
    for pid, lp in live.items():
        tp = tmpl.get(pid)
        if not tp:
            continue
        for fld in _PORT_FIELDS:
            lv = lp.get(fld)
            tv = tp.get(fld)
            if fld == "tags":
                lv = _normalize_tags(lv)
                tv = _normalize_tags(tv)
            if lv is not None and lv != tv:
                overrides.setdefault(pid, {})[fld] = lv
    return overrides

# ---- Template rebind / profile reassign "what-if" (read-only) ----
def list_org_templates(org_id: str) -> List[Dict[str, Any]]:
    try:
        return dashboard.organizations.getOrganizationConfigTemplates(org_id) or []
    except Exception:
        logging.exception("Failed to list config templates")
        return []

def get_template_profiles(org_id: str, template_id: Optional[str]) -> List[Dict[str, Any]]:
    if not template_id:
        return []
    try:
        return dashboard.switch.getOrganizationConfigTemplateSwitchProfiles(org_id, template_id) or []
    except Exception:
        logging.exception("Failed to read switch profiles for template %s", template_id)
        return []

def get_profile_ports(org_id: str, template_id: str, profile_id: str) -> List[Dict[str, Any]]:
    try:
        return dashboard.switch.getOrganizationConfigTemplateSwitchProfilePorts(org_id, template_id, profile_id) or []
    except Exception:
        logging.exception("Failed to read ports for profile %s (tpl %s)", profile_id, template_id)
        return []

def get_device_switch_ports(serial: str) -> List[Dict[str, Any]]:
    try:
        return dashboard.switch.getDeviceSwitchPorts(serial) or []
    except Exception:
        logging.exception("Failed to read device switch ports for %s", serial)
        return []

# --- Production-like template picking (regex heuristics)
def _pick_template_by_vlan_count_like_prod(templates: List[Dict[str, Any]], vlan_count: Optional[int]) -> Optional[Dict[str, Any]]:
    if vlan_count not in (3, 5):
        return None

    patterns: List[str] = []
    if vlan_count == 3:
        patterns = [r'NO\s*LEGACY.*MX*\b']
    elif vlan_count == 5:
        patterns = [r'3\s*X\s*DATA[_\s-]*VLAN.*MX75\b']

    for pat in patterns:
        rx = re.compile(pat, re.IGNORECASE)
        for t in templates:
            name = (t.get('name') or '')
            if rx.search(name):
                return t
    return None

def suggest_template(org_id: str, mx_models: List[str], vlan_count: int, templates: List[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    # 1) VLAN-count regex match like production
    by_vlan = _pick_template_by_vlan_count_like_prod(templates, vlan_count)
    if by_vlan:
        return by_vlan

    # 2) MX suffix fallback
    model_hint: Optional[str] = None
    for m in mx_models:
        if m.startswith("MX67"):
            model_hint = "MX67"
            break
        if m.startswith("MX75"):
            model_hint = "MX75"
            break
    if model_hint:
        for t in templates:
            name = (t.get("name") or "").strip().upper()
            if name.endswith(model_hint):
                return t

    # 3) Last resort
    return templates[0] if templates else None

def simulate_template_rebind_and_port_overrides(
    org_id: str,
    network_id: str,
    ms_list: List[Dict[str, Any]],
    mx_list: List[Dict[str, Any]],
    vlan_list: List[Dict[str, Any]],
) -> Tuple[Optional[str], Optional[str], List[Dict[str, Any]], List[Dict[str, Any]], List[Dict[str, Any]]]:
    """
    Returns: (sim_template_id, sim_template_name, planned_actions, port_override_rows, ms_without_profile_rows)
      - planned_actions: includes "would_bind_template" if currently unbound
                         and "would_assign_switch_profileId"/"would_reapply_port_overrides" entries.
      - port_override_rows: per-port overrides to reapply after simulated profile assign.
      - ms_without_profile_rows: MS devices with no compatible profile in the simulated template.
    """
    actions: List[Dict[str, Any]] = []
    port_rows: List[Dict[str, Any]] = []
    ms_wo_profile_rows: List[Dict[str, Any]] = []

    current_tpl_id = get_network_template_id(network_id)
    templates = list_org_templates(org_id)
    vlan_count = len(vlan_list)

    if current_tpl_id:
        # Already bound -> simulate staying on same template
        try:
            curr = dashboard.organizations.getOrganizationConfigTemplate(org_id, current_tpl_id)
            sim_tpl = curr
        except Exception:
            sim_tpl = {"id": current_tpl_id, "name": "(current template)"}
        sim_tpl_id = cast(str, sim_tpl.get("id"))
        sim_tpl_name = cast(str, sim_tpl.get("name", ""))
        actions.append({
            "scope": "Template",
            "action": "no_change_dry_run",
            "templateId": sim_tpl_id,
            "templateName": sim_tpl_name
        })
    else:
        # Unbound -> choose by production-like rules
        mx_models = [d.get("model", "") for d in mx_list]
        sim_tpl = suggest_template(org_id, mx_models, vlan_count, templates)
        if not sim_tpl:
            # Nothing to choose from; return gracefully
            return None, None, actions, port_rows, ms_wo_profile_rows
        sim_tpl_id = cast(str, sim_tpl.get("id"))
        sim_tpl_name = cast(str, sim_tpl.get("name", ""))
        actions.append({
            "scope": "Template",
            "action": "would_bind_template (simulation only)",
            "templateId": sim_tpl_id,
            "templateName": sim_tpl_name,
            "reason": f"Production-like heuristic (VLAN count={vlan_count}, MX hint from inventory)"
        })

    # Build profile maps for the simulated template (read-only)
    sim_profiles = get_template_profiles(org_id, sim_tpl_id)
    profile_name_to_id: Dict[str, str] = {p.get("name", ""): cast(str, p.get("switchProfileId")) for p in sim_profiles if p.get("switchProfileId")}
    model_to_profiles: Dict[str, List[Dict[str, Any]]] = {}
    for p in sim_profiles:
        for mdl in p.get("model", []):
            model_to_profiles.setdefault(mdl, []).append(p)

    # For each MS, simulate profile assignment and override reapply
    for sw in ms_list:
        serial = cast(str, sw.get("serial", ""))
        model = cast(str, sw.get("model", ""))
        name = cast(str, sw.get("name", "") or "")
        old_name = cast(str, sw.get("switchProfileName", "") or "")
        target_profile_id: Optional[str] = None
        target_profile_name: Optional[str] = None

        # Prefer same profile name if it exists in the simulated template
        if old_name and old_name in profile_name_to_id:
            target_profile_id = profile_name_to_id[old_name]
            target_profile_name = old_name
        else:
            # Else any profile that supports this MS model
            candidates = model_to_profiles.get(model, [])
            if candidates:
                target_profile_id = cast(str, candidates[0].get("switchProfileId"))
                target_profile_name = cast(str, candidates[0].get("name", ""))

        if not target_profile_id:
            actions.append({
                "scope": "MS",
                "serial": serial,
                "model": model,
                "action": "no_suitable_switch_profile_found_in_sim_template"
            })
            ms_wo_profile_rows.append({
                "serial": serial,
                "model": model,
                "deviceName": name,
                "previousProfileName": old_name,
                "templateId": sim_tpl_id,
                "templateName": sim_tpl_name
            })
            continue

        actions.append({
            "scope": "MS",
            "serial": serial,
            "action": "would_assign_switch_profileId (simulation only)",
            "targetProfileId": target_profile_id,
            "targetProfileName": target_profile_name,
            "templateId": sim_tpl_id,
            "templateName": sim_tpl_name
        })

        # Compute port overrides vs target profile (read-only)
        live_ports = get_device_switch_ports(serial)
        tmpl_ports = get_profile_ports(org_id, sim_tpl_id, target_profile_id)
        overrides = compute_port_overrides(live_ports, tmpl_ports)

        actions.append({
            "scope": "MS",
            "serial": serial,
            "action": "would_reapply_port_overrides (simulation only)",
            "override_port_count": len(overrides),
        })

        for port_id, patch in overrides.items():
            for fld, val in patch.items():
                port_rows.append({
                    "serial": serial,
                    "portId": str(port_id),
                    "field": fld,
                    "value": val if not isinstance(val, (dict, list)) else json.dumps(val, ensure_ascii=False),
                    "profileName": target_profile_name or "",
                    "profileId": target_profile_id,
                    "templateId": sim_tpl_id,
                })

    return sim_tpl_id, sim_tpl_name, actions, port_rows, ms_wo_profile_rows

# ---- High-level "what would happen"
def plan_high_level_changes(mx: List[Dict[str, Any]], mr: List[Dict[str, Any]], warm_spare: Dict[str, Any]) -> List[Dict[str, Any]]:
    actions: List[Dict[str, Any]] = []

    has_mx64 = any((d.get("model") or "").startswith("MX64") for d in mx)
    has_newer_mx = any((d.get("model") or "").startswith(("MX67", "MX75")) for d in mx)
    if has_mx64:
        actions.append({
            "scope": "MX",
            "action": "would_remove_mx64_after_upgrade",
            "precondition": "Newer MX present/claimed (not performed in dry-run)",
            "detected_newer_mx_in_network": has_newer_mx
        })

    has_mr33 = any((d.get("model") or "") == "MR33" for d in mr)
    has_other_wireless = any((d.get("model") or "").startswith(("MR", "CW")) and (d.get("model") != "MR33") for d in mr)
    if has_mr33 and has_other_wireless:
        actions.append({
            "scope": "MR",
            "action": "would_retire_legacy_mr33",
            "precondition": "Replacement APs present/claimed (not performed in dry-run)"
        })

    if warm_spare:
        actions.append({
            "scope": "WarmSpare",
            "action": "no_change_dry_run",
            "enabled": bool(warm_spare.get("enabled")),
            "primarySerial": warm_spare.get("primarySerial"),
            "secondarySerial": warm_spare.get("spareSerial")
        })

    return actions

# ---------------- Excel Exports ----------------
def _auto_size(ws: Worksheet) -> None:
    max_col: int = ws.max_column
    max_row: int = ws.max_row
    for c in range(1, max_col + 1):
        max_len = 0
        for r in range(1, max_row + 1):
            v = ws.cell(row=r, column=c).value
            s = "" if v is None else str(v)
            if len(s) > max_len:
                max_len = len(s)
        ws.column_dimensions[get_column_letter(c)].width = min(max_len + 2, 60)

def export_snapshot(
    path: str,
    network_id: str,
    network_name: str,
    vlan_list: List[Dict[str, Any]],
    mx_list: List[Dict[str, Any]],
    ms_list: List[Dict[str, Any]],
    mr_list: List[Dict[str, Any]],
    planned_actions: Optional[List[Dict[str, Any]]] = None,
    port_override_rows: Optional[List[Dict[str, Any]]] = None,
    ms_wo_profile_rows: Optional[List[Dict[str, Any]]] = None,
    title: str = "Snapshot",
) -> None:
    wb: Workbook = Workbook()
    ws: Worksheet = cast(Worksheet, wb.active)
    ws.title = title

    # VLANs
    ws.append(["section", "network_id", "network_name", "item_type", "id", "name", "subnet", "applianceIp", "dhcpHandling", "extra"])
    for v in vlan_list:
        extra = {k: v.get(k) for k in ("fixedIpAssignments", "reservedIpRanges") if k in v}
        ws.append([
            "vlans", network_id, network_name, "vlan",
            str(v.get("id", "")),
            str(v.get("name", "") or ""),
            str(v.get("subnet", "") or ""),
            str(v.get("applianceIp", "") or ""),
            str(v.get("dhcpHandling", "") or ""),
            json.dumps(extra, ensure_ascii=False) if extra else ""
        ])

    # Devices
    ws.append([]); ws.append(["section", "network_id", "network_name", "item_type", "serial", "model", "name", "address", "tags", "switchProfileId", "switchProfileName"])
    for d in (mx_list + ms_list + mr_list):
        tags_list: List[str] = d.get("tags", []) if isinstance(d.get("tags", []), list) else []
        ws.append([
            "devices", network_id, network_name, "device",
            str(d.get("serial", "") or ""),
            str(d.get("model", "") or ""),
            str(d.get("name", "") or ""),
            str(d.get("address", "") or ""),
            " ".join(tags_list),
            str(d.get("switchProfileId", "") or ""),
            str(d.get("switchProfileName", "") or "")
        ])

    # Planned actions
    if planned_actions is not None:
        ws2 = wb.create_sheet("Planned_Actions")
        ws2.append(["scope", "action", "details"])
        for a in planned_actions:
            ws2.append([str(a.get("scope", "")), str(a.get("action", "")), json.dumps(a, ensure_ascii=False)])
        _auto_size(ws2)

    # Port override simulation (per-port)
    if port_override_rows:
        ws3 = wb.create_sheet("Port_Overrides_Sim")
        ws3.append(["serial", "portId", "field", "value", "profileName", "profileId", "templateId"])
        for r in port_override_rows:
            ws3.append([
                r.get("serial", ""),
                r.get("portId", ""),
                r.get("field", ""),
                r.get("value", ""),
                r.get("profileName", ""),
                r.get("profileId", ""),
                r.get("templateId", "")
            ])
        _auto_size(ws3)

    # MS devices with no compatible profile (triage)
    if ms_wo_profile_rows:
        ws4 = wb.create_sheet("MS_No_Compatible_Profile")
        ws4.append(["serial", "model", "deviceName", "previousProfileName", "templateId", "templateName"])
        for r in ms_wo_profile_rows:
            ws4.append([
                r.get("serial", ""),
                r.get("model", ""),
                r.get("deviceName", ""),
                r.get("previousProfileName", ""),
                r.get("templateId", ""),
                r.get("templateName", "")
            ])
        _auto_size(ws4)

    _auto_size(ws)
    wb.save(path)
    print(f"📄 Exported: {path}")

def export_diff(
    path: str,
    pre_vlans: List[Dict[str, Any]],
    post_vlans: List[Dict[str, Any]],
) -> None:
    def key(v: Dict[str, Any]) -> str:
        return str(v.get("id", ""))

    pre_map = {key(v): v for v in pre_vlans}
    post_map = {key(v): v for v in post_vlans}

    wb = Workbook()
    ws = cast(Worksheet, wb.active)
    ws.title = "Diff"

    ws.append(["vlanId", "field", "pre", "post", "note"])

    for vid, v_pre in pre_map.items():
        v_post = post_map.get(vid, {})
        pre_dhcp = v_pre.get("dhcpHandling", "")
        post_dhcp = v_post.get("dhcpHandling", "")
        note = ""
        if pre_dhcp != post_dhcp:
            note = "Changed by simulation (unlikely unless edited upstream)"
        ws.append([vid, "dhcpHandling", str(pre_dhcp), str(post_dhcp), note])

        pre_has_fixed = "fixedIpAssignments" in v_pre and bool(v_pre.get("fixedIpAssignments"))
        pre_has_reserved = "reservedIpRanges" in v_pre and bool(v_pre.get("reservedIpRanges"))
        post_has_fixed = "fixedIpAssignments" in v_post and bool(v_post.get("fixedIpAssignments"))
        post_has_reserved = "reservedIpRanges" in v_post and bool(v_post.get("reservedIpRanges"))

        if pre_has_fixed != post_has_fixed:
            ws.append([vid, "fixedIpAssignments_present", str(pre_has_fixed), str(post_has_fixed), "Stripped in simulation if DHCP ≠ server"])
        if pre_has_reserved != post_has_reserved:
            ws.append([vid, "reservedIpRanges_present", str(pre_has_reserved), str(post_has_reserved), "Stripped in simulation if DHCP ≠ server"])

    _auto_size(ws)
    wb.save(path)
    print(f"📄 Exported: {path}")

# ---------------- Main (dry-run only) ----------------
if __name__ == "__main__":
    print("=== DRY RUN: Snapshot + Simulated Plan (NO CHANGES MADE) ===")
    org_id = select_org()
    net_id, net_name = select_network_interactive(org_id)

    # Live reads
    mx, ms, mr = fetch_devices(net_id)
    vlans_pre = fetch_vlan_details(net_id)
    warm_spare = get_warm_spare_status(net_id)

    print(f"\nNetwork: {net_name} ({net_id})")
    print(f"Devices: MX={len(mx)}  MS={len(ms)}  MR={len(mr)}")
    print(f"VLANs:   {len(vlans_pre)}")
    if warm_spare:
        print(f"Warm spare enabled: {bool(warm_spare.get('enabled'))}")

    # Build the in-memory VLAN plan (NO changes)
    vlans_post_sim, vlan_actions = simulate_vlan_payloads(vlans_pre)

    # Template rebind "what-if" + simulated profile reassign + port overrides + triage
    sim_tpl_id, sim_tpl_name, tpl_actions, port_override_rows, ms_wo_profile_rows = simulate_template_rebind_and_port_overrides(
        org_id, net_id, ms_list=ms, mx_list=mx, vlan_list=vlans_pre
    )

    # High-level notes (MX/MR/Warm spare)
    hi_actions = plan_high_level_changes(mx, mr, warm_spare)

    planned_actions = vlan_actions + tpl_actions + hi_actions

    # Exports
    base = _slug_filename(net_name) or "network"
    pre_path = f"{base}_pre_{timestamp}.xlsx"
    post_path = f"{base}_post_simulated_{timestamp}.xlsx"
    diff_path = f"{base}_diff_{timestamp}.xlsx"

    export_snapshot(
        pre_path, net_id, net_name, vlans_pre, mx, ms, mr,
        planned_actions=None,
        port_override_rows=None,
        ms_wo_profile_rows=None,
        title="PRE_Snapshot"
    )

    export_snapshot(
        post_path, net_id, net_name, vlans_post_sim, mx, ms, mr,
        planned_actions=planned_actions,
        port_override_rows=port_override_rows,
        ms_wo_profile_rows=ms_wo_profile_rows,
        title="POST_Simulated"
    )

    export_diff(diff_path, vlans_pre, vlans_post_sim)

    print("\n✅ Dry run complete — no changes made to Meraki Dashboard.")
    print("   Generated files:")
    print(f"   - {pre_path}")
    print(f"   - {post_path}")
    print(f"   - {diff_path}")

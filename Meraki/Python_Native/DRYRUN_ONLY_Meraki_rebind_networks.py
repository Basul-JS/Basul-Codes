
"""
Meraki Dry-Run Network Inspector (Pylance-safe)

- Read-only: never claims devices, never renames, never PUT/POST to Meraki.
- Prints/logs: networks, devices, VLANs, and "would-be" VLAN payloads.
- Robust network search: pagination + unicode-normalized matching.
- Wireless recognition: MR* and CW* models are treated as APs (e.g., CW9172I).
- Pylance-compliant typing with safe_int/safe_str guards.

Requirements:
  pip install openpyxl requests
"""

from __future__ import annotations

import csv
import json
import logging
import os
import re
import signal
import sys
import time
import unicodedata
from datetime import datetime
from getpass import getpass
from typing import Any, Dict, List, Optional, Tuple, Set

import requests
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# =====================
# Config & Constants
# =====================

REQUEST_TIMEOUT: int = 30
BASE_URL: str = "https://api.meraki.com/api/v1"
MAX_RETRIES: int = 5

EXCLUDED_VLANS: Set[int] = {100, 110, 210, 220, 230, 235, 240}
WIRELESS_PREFIXES: Tuple[str, ...] = ("MR", "CW")

# Hard-enforced DRY RUN for this file
DRY_RUN: bool = True

timestamp: str = datetime.now().strftime("%Y%m%d_%H%M%S")
logging.basicConfig(
    filename=f"meraki_dryrun_{timestamp}.log",
    level=logging.DEBUG,
    format="%(asctime)s - %(levelname)s - %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
CSV_LOGFILE: str = f"meraki_dryrun_audit_{timestamp}.csv"

# =====================
# CSV audit (read-only notes)
# =====================

def log_change(
    event: str,
    details: str,
    *,
    username: Optional[str] = None,
    org_id: Optional[str] = None,
    org_name: Optional[str] = None,
    network_id: Optional[str] = None,
    network_name: Optional[str] = None,
    misc: Optional[str] = None,
) -> None:
    file_exists = os.path.isfile(CSV_LOGFILE)
    with open(CSV_LOGFILE, mode="a", newline="", encoding="utf-8") as csvfile:
        writer = csv.writer(csvfile)
        if not file_exists:
            writer.writerow(
                [
                    "timestamp",
                    "event",
                    "details",
                    "user",
                    "org_id",
                    "org_name",
                    "network_id",
                    "network_name",
                    "misc",
                ]
            )
        writer.writerow(
            [
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                event,
                details,
                username or "",
                org_id or "",
                org_name or "",
                network_id or "",
                network_name or "",
                misc or "",
            ]
        )

# =====================
# Prompts (minimal)
# =====================

OPERATOR: str = input("Enter your name or initials for audit logs: ").strip() or "operator"
print("DRY RUN MODE: No changes will be made to the Meraki org.")

# =====================
# API auth
# =====================

def validate_api_key(key: str) -> bool:
    return bool(re.fullmatch(r"[A-Fa-f0-9]{40}", key or ""))

MAX_API_KEY_ATTEMPTS: int = 4
attempts = 0
API_KEY: Optional[str] = None
while attempts < MAX_API_KEY_ATTEMPTS:
    API_KEY = getpass("Enter your Meraki API key (hidden): ").strip()
    if validate_api_key(API_KEY):
        break
    attempts += 1
    print(f"❌ Invalid API key. ({MAX_API_KEY_ATTEMPTS - attempts} attempt(s) left)")
else:
    print("❌ Maximum attempts reached. Exiting.")
    raise SystemExit(1)

HEADERS: Dict[str, str] = {
    "X-Cisco-Meraki-API-Key": API_KEY or "",
    "Content-Type": "application/json",
    "Accept": "application/json",
}

# Graceful abort
_aborted: bool = False
def _handle_sigint(signum: int, frame: Any) -> None:  # noqa: ARG001
    global _aborted
    _aborted = True
    print("\nReceived Ctrl+C — attempting graceful shutdown...")
    log_change("workflow_abort", "User interrupted with SIGINT", username=OPERATOR)
signal.signal(signal.SIGINT, _handle_sigint)

# =====================
# Safe converters (Pylance helpers)
# =====================

def safe_str(x: Any) -> str:
    """Return a string for logging/keys without ever passing None through."""
    return "" if x is None else str(x)

def safe_int(x: Any) -> Optional[int]:
    """Return an int if x looks like an int, else None (avoids passing None to int())."""
    if x is None:
        return None
    s = str(x).strip()
    # allow pure digits; if Meraki ever sends '41' as string, it's fine
    return int(s) if s.isdigit() else None

# =====================
# HTTP layer (GET only used)
# =====================

class MerakiAPIError(Exception):
    def __init__(self, status_code: int, text: str, json_body: Optional[Any], url: str):
        super().__init__(f"Meraki API error: {status_code} {text}")
        self.status_code = status_code
        self.text = text
        self.json_body = json_body
        self.url = url

def _request(method: str, path: str, *, params: Optional[Dict[str, Any]] = None, json_data: Optional[Any] = None) -> Any:
    url = f"{BASE_URL}{path}"
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            if method == "GET":
                resp = requests.get(url, headers=HEADERS, params=params, timeout=REQUEST_TIMEOUT)
            else:
                # We never call write methods in this dry-run file
                raise ValueError("Only GET is allowed in dry-run script.")
            if resp.status_code == 429:
                ra = resp.headers.get("Retry-After")
                wait: float
                if ra is not None:
                    try:
                        wait = max(0.5, min(float(ra), 30.0))
                    except ValueError:
                        wait = min(2 ** (attempt - 1), 30.0)
                else:
                    wait = min(2 ** (attempt - 1), 30.0)
                logging.warning("429 rate limit for %s. Sleeping %.1fs and retrying...", url, wait)
                time.sleep(wait)
                continue
            if not resp.ok:
                try:
                    body = resp.json()
                except Exception:
                    body = None
                logging.error("%s %s -> %s %s", method, url, resp.status_code, resp.text)
                raise MerakiAPIError(resp.status_code, resp.text, body, url)
            if resp.text:
                try:
                    return resp.json()
                except Exception:
                    return resp.text
            return None
        except MerakiAPIError:
            raise
        except Exception as e:
            if attempt == MAX_RETRIES:
                logging.exception("HTTP error for %s: %s", url, e)
                raise
            wait = min(2 ** attempt, 30.0)
            logging.warning("HTTP exception %s for %s. Retrying in %.1fs...", e, url, wait)
            time.sleep(wait)

def meraki_get(path: str, params: Optional[Dict[str, Any]] = None) -> Any:
    return _request("GET", path, params=params)

# =====================
# Helpers (search, typing-safe)
# =====================

def _norm(s: Optional[str]) -> str:
    """
    Normalize a string for robust matching:
    - unicode NFKC
    - convert en/em dashes to hyphen
    - collapse whitespace
    - casefold for case-insensitive compare
    """
    base: str = s or ""
    base = unicodedata.normalize("NFKC", base)
    base = base.replace("–", "-").replace("—", "-")
    base = re.sub(r"\s+", " ", base).strip()
    return base.casefold()

def meraki_list_networks_all(org_id: str) -> List[Dict[str, Any]]:
    """
    Returns ALL networks in an org using Meraki's cursor pagination.
    """
    all_nets: List[Dict[str, Any]] = []
    per_page: int = 1000
    starting_after: Optional[str] = None

    while True:
        params: Dict[str, Any] = {"perPage": per_page}
        if starting_after:
            params["startingAfter"] = starting_after
        page_raw: Any = meraki_get(f"/organizations/{org_id}/networks", params=params)
        page: List[Dict[str, Any]] = page_raw if isinstance(page_raw, list) else []
        if not page:
            break
        all_nets.extend(page)
        if len(page) < per_page:
            break
        last = page[-1]
        starting_after = safe_str(last.get("id"))
        if not starting_after:
            break
    return all_nets

def fetch_matching_networks(org_id: str, partial: str) -> List[Dict[str, Any]]:
    partial_n: str = _norm(partial)
    nets: List[Dict[str, Any]] = meraki_list_networks_all(org_id)
    matches: List[Dict[str, Any]] = []
    for n in nets:
        name = _norm(safe_str(n.get("name")))
        if partial_n in name:
            matches.append(n)
    logging.debug("Found %d networks matching '%s' (normalized)", len(matches), partial)
    return matches

def _is_wireless_model(model: Optional[str]) -> bool:
    return bool(model) and str(model).upper().startswith(WIRELESS_PREFIXES)

# VLAN helpers (read/build-only)

def is_vlans_disabled_error(exc: Exception) -> bool:
    needle = "VLANs are not enabled for this network"
    try:
        if isinstance(exc, MerakiAPIError):
            if exc.status_code == 400:
                if exc.json_body and isinstance(exc.json_body, dict):
                    errs = exc.json_body.get("errors")
                    if errs and any(needle in str(e) for e in errs):
                        return True
                if needle in (exc.text or ""):
                    return True
        return needle in str(exc)
    except Exception:
        return False

def fetch_vlan_details(network_id: str) -> List[Dict[str, Any]]:
    try:
        vlans_raw: Any = meraki_get(f"/networks/{network_id}/appliance/vlans")
        vlans: List[Dict[str, Any]] = vlans_raw if isinstance(vlans_raw, list) else []
        filtered: List[Dict[str, Any]] = []
        for v in vlans:
            vid = safe_int(v.get("id"))
            if vid is None:
                continue
            if vid in EXCLUDED_VLANS:
                continue
            filtered.append(v)
        logging.debug("Fetched VLANs: %d (excluded %d)", len(filtered), len(vlans) - len(filtered))
        return filtered
    except MerakiAPIError as e:
        if is_vlans_disabled_error(e):
            logging.warning("VLAN endpoints unavailable (VLANs disabled). Returning empty list.")
            return []
        logging.exception("Failed to fetch VLANs")
        return []
    except Exception:
        logging.exception("Failed to fetch VLANs")
        return []

def _dhcp_mode(val: Optional[str]) -> str:
    v = (val or "").strip().lower()
    if v in {"run a dhcp server", "run dhcp server", "server", "enabled", "on"}:
        return "server"
    if "relay" in v:
        return "relay"
    if v in {"do not respond", "do not respond to dhcp requests", "off", "disabled", "none"}:
        return "off"
    return "off"

def build_vlan_update_payloads(vlan_list: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    """
    Pure builder: returns a dict of {vlan_id: payload} of what we WOULD send,
    respecting DHCP mode rules (no illegal keys).
    """
    out: Dict[str, Dict[str, Any]] = {}
    for v in vlan_list:
        vid = safe_int(v.get("id"))
        vlan_id: str = safe_str(vid)
        if not vlan_id:
            # skip if VLAN id isn't numeric
            continue

        payload: Dict[str, Any] = {}

        if v.get("applianceIp"):
            payload["applianceIp"] = v.get("applianceIp")
        if v.get("subnet"):
            payload["subnet"] = v.get("subnet")

        dhcp_handling_raw = v.get("dhcpHandling")
        if dhcp_handling_raw:
            payload["dhcpHandling"] = dhcp_handling_raw

        mode = _dhcp_mode(dhcp_handling_raw)

        if mode == "server":
            if v.get("fixedIpAssignments"):
                payload["fixedIpAssignments"] = v.get("fixedIpAssignments")
            if v.get("reservedIpRanges"):
                payload["reservedIpRanges"] = v.get("reservedIpRanges")
            if v.get("dnsNameservers"):
                payload["dnsNameservers"] = v.get("dnsNameservers")
        elif mode == "relay":
            relay_ips = v.get("dhcpRelayServerIps") or v.get("dhcpRelayServerIp")
            if relay_ips:
                payload["dhcpRelayServerIps"] = relay_ips
        # mode off: no extras

        out[vlan_id] = payload
    return out

# =====================
# Snapshot export (read-only)
# =====================

def _slug_filename(s: str) -> str:
    s2 = re.sub(r"[^A-Za-z0-9._-]+", "-", s).strip("-_")
    return s2[:80]

def _network_tag_from_name(name: str) -> str:
    parts = name.split("-")
    if len(parts) >= 2 and parts[1].isdigit():
        return f"{parts[0]}-{parts[1]}"
    return name

def export_network_snapshot_xlsx(
    org_id: str,
    network_id: str,
    network_name: str,
    template_id: Optional[str],
    vlan_list: List[Dict[str, Any]],
    mx_list: List[Dict[str, Any]],
    ms_list: List[Dict[str, Any]],
    mr_list: List[Dict[str, Any]],
    outfile: Optional[str] = None,
) -> None:
    def _json(x: Any) -> str:
        try:
            return json.dumps(x, ensure_ascii=False)
        except Exception:
            return str(x)

    out_path: str
    if outfile:
        out_path = outfile
    else:
        base = _network_tag_from_name(network_name)
        out_path = f"{_slug_filename(base)}_{timestamp}.xlsx"

    wb: Workbook = Workbook()
    ws: Worksheet = wb.active  # type: ignore[assignment]
    ws.title = "Snapshot"

    header: List[str] = [
        "section",
        "network_id",
        "network_name",
        "item_type",
        "col1",
        "col2",
        "col3",
        "col4",
        "col5",
        "extra_info",
    ]
    ws.append(header)

    # template name (best effort)
    tpl_name: str = ""
    if template_id:
        try:
            tpl_raw: Any = meraki_get(f"/organizations/{org_id}/configTemplates/{template_id}")
            tpl: Dict[str, Any] = tpl_raw if isinstance(tpl_raw, dict) else {}
            tpl_name = str(tpl.get("name", "") or "")
        except Exception:
            logging.exception("Could not fetch template name for snapshot")

    ws.append(
        [
            "template",
            network_id,
            network_name,
            "template",
            template_id or "",
            tpl_name,
            "",
            "",
            "",
            "",
        ]
    )

    for v in vlan_list:
        ws.append(
            [
                "vlans",
                network_id,
                network_name,
                "vlan",
                safe_str(safe_int(v.get("id"))),
                safe_str(v.get("name")),
                safe_str(v.get("subnet")),
                safe_str(v.get("applianceIp")),
                safe_str(v.get("dhcpHandling")),
                _json({k: v.get(k) for k in v.keys() - {"id", "name", "subnet", "applianceIp", "dhcpHandling"}}),
            ]
        )

    def _device_row(d: Dict[str, Any]) -> List[str]:
        tags_val = d.get("tags", [])
        if isinstance(tags_val, list):
            tags_list = [str(t) for t in tags_val]
        else:
            tags_list = [t for t in str(tags_val or "").split() if t]
        return [
            "devices",
            network_id,
            network_name,
            "device",
            safe_str(d.get("serial")),
            safe_str(d.get("model")),
            safe_str(d.get("name")),
            safe_str(d.get("address")),
            " ".join(tags_list),
            "",
        ]

    for d in (mx_list + ms_list + mr_list):
        ws.append(_device_row(d))

    # autosize
    max_col: int = ws.max_column
    max_row: int = ws.max_row
    for col_idx in range(1, max_col + 1):
        max_len = 0
        for row_idx in range(1, max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                s = str(val)
                if len(s) > max_len:
                    max_len = len(s)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

    wb.save(out_path)
    print(f"📄 Snapshot exported to Excel: {out_path}")
    log_change("snapshot_export", f"Exported network snapshot to {out_path}",
               username=OPERATOR, network_id=network_id, network_name=network_name)

# =====================
# Org & Network selection (read-only)
# =====================

def select_org() -> str:
    orgs_raw: Any = meraki_get("/organizations")
    orgs: List[Dict[str, Any]] = orgs_raw if isinstance(orgs_raw, list) else []
    if not orgs:
        print("\n❌ No Organisations returned from API -----------")
        print("   Please retry when Org is known *******")
        sys.exit(1)

    print("Organizations:")
    for idx, org in enumerate(orgs, 1):
        print(f"{idx}. {safe_str(org.get('name'))} (ID: {safe_str(org.get('id'))})")

    raw = input("Select organization by number (or press Enter to cancel): ").strip()
    if not raw:
        print("\n❌ No Organisation selected -----------")
        sys.exit(1)

    try:
        org_idx = int(raw)
        if org_idx < 1 or org_idx > len(orgs):
            raise ValueError("out of range")
    except Exception:
        print("\n❌ Invalid Organisation selection -----------")
        sys.exit(1)

    return safe_str(orgs[org_idx - 1].get("id"))

def select_network_interactive(org_id: str) -> Tuple[str, str]:
    while True:
        partial = input("Enter partial network name to search (or press Enter to cancel): ").strip()
        if not partial:
            print("\n❌ No Network selected -----------")
            sys.exit(1)

        networks = fetch_matching_networks(org_id, partial)
        if not networks:
            print("\n❌ No matching networks found -----------")
            retry = input("Search again? (y/N): ").strip().lower()
            if retry != "y":
                print("\n❌ No Network selected -----------")
                sys.exit(1)
            continue

        if len(networks) == 1:
            only = networks[0]
            name = safe_str(only.get("name"))
            nid = safe_str(only.get("id"))
            print(f"\n1 match: {name} (ID: {nid})")
            confirm = input("Use this network? (Y/n): ").strip().lower()
            if confirm in {"", "y", "yes"}:
                print(f"Selected network: {name} (ID: {nid})")
                return nid, name
            else:
                continue

        print("\nMultiple networks found:")
        for idx, net in enumerate(networks, 1):
            print(f"{idx}. {safe_str(net.get('name'))} (ID: {safe_str(net.get('id'))})")

        while True:
            raw = input("Select the network by number (or press Enter to cancel): ").strip()
            if not raw:
                print("\n❌ No Network selected -----------")
                sys.exit(1)
            if raw.isdigit():
                choice = int(raw)
                if 1 <= choice <= len(networks):
                    chosen = networks[choice - 1]
                    name = safe_str(chosen.get("name"))
                    nid = safe_str(chosen.get("id"))
                    print(f"Selected network #{choice}: {name} (ID: {nid})")
                    return nid, name
            print("❌ Invalid selection. Please enter a valid number from the list.")

# =====================
# Device fetcher (read-only, wireless-aware)
# =====================

def fetch_devices(org_id: str, network_id: str) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], List[Dict[str, Any]]]:
    devs_raw: Any = meraki_get(f"/networks/{network_id}/devices")
    devs: List[Dict[str, Any]] = devs_raw if isinstance(devs_raw, list) else []

    def _mk(d: Dict[str, Any]) -> Dict[str, Any]:
        tags_val = d.get("tags", [])
        if isinstance(tags_val, list):
            tags_list = [str(t) for t in tags_val]
        else:
            tags_list = [t for t in str(tags_val or "").split() if t]
        return {
            "serial": safe_str(d.get("serial")),
            "model": safe_str(d.get("model")),
            "tags": tags_list,
            "address": safe_str(d.get("address")),
            "name": safe_str(d.get("name")),
        }

    mx: List[Dict[str, Any]] = [_mk(d) for d in devs if safe_str(d.get("model")).upper().startswith("MX")]
    ms: List[Dict[str, Any]] = [_mk(d) for d in devs if safe_str(d.get("model")).upper().startswith("MS")]

    mr: List[Dict[str, Any]] = []
    for d in devs:
        if _is_wireless_model(safe_str(d.get("model"))):
            mr.append(_mk(d))

    logging.debug("Fetched devices: MX=%d, MS=%d, MR/CW=%d", len(mx), len(ms), len(mr))
    return mx, ms, mr

# =====================
# DRY-RUN processing
# =====================

def dryrun_process_network(org_id: str, network_id: str, network_name: str) -> None:
    print(f"\n=== DRY RUN: {network_name} ({network_id}) ===")
    log_change("dryrun_network_start", f"Processing {network_name}",
               username=OPERATOR, network_id=network_id, network_name=network_name)

    # Basic network info
    try:
        net_info_raw: Any = meraki_get(f"/networks/{network_id}")
        net_info: Dict[str, Any] = net_info_raw if isinstance(net_info_raw, dict) else {}
        tpl_id = net_info.get("configTemplateId")
        print(f"Template bound: {safe_str(tpl_id) or 'None'}")
    except Exception:
        logging.exception("Failed to read network info")
        tpl_id = None

    # Devices
    mx_list, ms_list, mr_list = fetch_devices(org_id, network_id)
    print(f"Devices — MX: {len(mx_list)}  MS: {len(ms_list)}  MR/CW: {len(mr_list)}")
    for d in (mx_list + ms_list + mr_list):
        print(f" - {d.get('model','')} {d.get('serial','')}  name='{d.get('name','')}'")

    # VLANs
    vlans = fetch_vlan_details(network_id)
    print(f"VLANs (excluding {sorted(EXCLUDED_VLANS)}): {len(vlans)}")
    for v in vlans:
        vid = safe_int(v.get("id"))
        print(f" - VLAN {safe_str(vid)}  {safe_str(v.get('name'))}  {safe_str(v.get('subnet'))}  DHCP={safe_str(v.get('dhcpHandling'))}")

    # Build (do not send) VLAN payloads
    payloads = build_vlan_update_payloads(vlans)
    print("\n[DRY RUN] VLAN payloads that would be sent (but are NOT sent):")
    for vid_s, payload in payloads.items():
        print(f" * VLAN {vid_s}: {json.dumps(payload, indent=2)}")
        logging.info("[DRY RUN] Would update VLAN %s with payload: %s", vid_s, json.dumps(payload))

    # Optional: export snapshot XLSX
    export = input("\nExport a read-only XLSX snapshot? (y/N): ").strip().lower()
    if export == "y":
        export_network_snapshot_xlsx(
            org_id=org_id,
            network_id=network_id,
            network_name=network_name,
            template_id=safe_str(tpl_id) if tpl_id else None,
            vlan_list=vlans,
            mx_list=mx_list,
            ms_list=ms_list,
            mr_list=mr_list,
            outfile=f"{_slug_filename(_network_tag_from_name(network_name))}_dryrun_{timestamp}.xlsx",
        )

    log_change("dryrun_network_end", f"Finished {network_name}",
               username=OPERATOR, network_id=network_id, network_name=network_name)

# =====================
# Main
# =====================

def main() -> None:
    log_change("workflow_start", "Dry-run script started", username=OPERATOR)

    # -------- Select Org --------
    org_id = select_org()

    # -------- Select Network(s) --------
    while True:
        network_id, network_name = select_network_interactive(org_id)
        dryrun_process_network(org_id, network_id, network_name)

        if _aborted:
            break
        again = input("\nProcess another network? (y/N): ").strip().lower()
        if again != "y":
            break

    log_change("workflow_end", "Dry-run script finished", username=OPERATOR)
    print("\nDone. No changes were made.")

if __name__ == "__main__":
    main()
